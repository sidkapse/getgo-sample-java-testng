"""
generate_test_data.py
Generates test-data/getgo_test_data.xlsx with sample test data rows.
Run once: python generate_test_data.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
CELL_FONT    = Font(name="Calibri", size=10)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

def styled_header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 22

def add_rows(ws, rows):
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font      = CELL_FONT
            cell.alignment = ALIGN_CENTER

def set_col_widths(ws, widths):
    for col, width in zip(ws.column_dimensions, widths):
        ws.column_dimensions[col].width = width

wb = openpyxl.Workbook()

# ── Sheet 1: CreateBooking ─────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "CreateBooking"
styled_header(ws1, [
    "passenger_id", "pickup_lat", "pickup_lng",
    "dropoff_lat",  "dropoff_lng", "vehicle_type", "payment_method"
])
add_rows(ws1, [
    ["PSG-1001", 1.3048, 103.8318, 1.2966, 103.8536, "STANDARD", "CARD"],
    ["PSG-1002", 1.2835, 103.8607, 1.3190, 103.8412, "PREMIUM",  "WALLET"],
    ["PSG-1003", 1.3521, 103.8198, 1.3000, 103.7800, "XL",       "CASH"],
    ["PSG-1004", 1.2950, 103.8550, 1.3100, 103.8200, "STANDARD", "CARD"],
    ["PSG-1005", 1.3200, 103.8400, 1.2800, 103.8600, "PREMIUM",  "WALLET"],
])
set_col_widths(ws1, ["A","B","C","D","E","F","G"])
ws1.column_dimensions["A"].width = 14
ws1.column_dimensions["B"].width = 12
ws1.column_dimensions["C"].width = 12
ws1.column_dimensions["D"].width = 12
ws1.column_dimensions["E"].width = 12
ws1.column_dimensions["F"].width = 14
ws1.column_dimensions["G"].width = 16

# ── Sheet 2: RegisterDriver ────────────────────────────────────────────────
ws2 = wb.create_sheet("RegisterDriver")
styled_header(ws2, [
    "full_name", "nric", "phone",
    "vehicle_plate", "vehicle_type", "license_expiry"
])
add_rows(ws2, [
    ["Ahmad Bin Ali",     "S9012345A", "+6591234567", "SBX1234Z", "STANDARD", "2026-12-31"],
    ["Tan Wei Ming",      "T8523456B", "+6598765432", "SGX5678Y", "PREMIUM",  "2025-06-30"],
    ["Kumar Rajan",       "S7834567C", "+6587654321", "SHY9012A", "XL",       "2027-03-15"],
    ["Siti Binte Rahmat", "T9145678D", "+6592345678", "SJK3456B", "STANDARD", "2026-09-01"],
])
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 14
ws2.column_dimensions["F"].width = 16

# ── Sheet 3: RideLifecycle ─────────────────────────────────────────────────
ws3 = wb.create_sheet("RideLifecycle")
styled_header(ws3, [
    "booking_id", "driver_id", "start_lat", "start_lng",
    "end_lat", "end_lng", "distance_km"
])
add_rows(ws3, [
    ["BK-9001", "DR-5001", 1.3048, 103.8318, 1.2966, 103.8536, 4.2],
    ["BK-9002", "DR-5002", 1.2835, 103.8607, 1.3190, 103.8412, 6.8],
    ["BK-9003", "DR-5003", 1.3521, 103.8198, 1.3000, 103.7800, 9.1],
])
for col in ["A","B","C","D","E","F","G"]:
    ws3.column_dimensions[col].width = 14

os.makedirs("test-data", exist_ok=True)
output = "test-data/getgo_test_data.xlsx"
wb.save(output)
print(f"Generated: {output}")
print(f"Sheets: {wb.sheetnames}")
